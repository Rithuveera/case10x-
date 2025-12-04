from flask import Flask, request, jsonify, render_template, send_file
import google.generativeai as genai
import os
import json
import sqlite3
from datetime import datetime
from werkzeug.utils import secure_filename
import PyPDF2
from docx import Document
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'docx', 'txt'}

# Configure Gemini AI
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    print("WARNING: GEMINI_API_KEY not found in environment variables!")
else:
    genai.configure(api_key=GEMINI_API_KEY)

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database initialization
def init_db():
    conn = sqlite3.connect('database.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS test_cases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requirement_file TEXT NOT NULL,
            test_case_name TEXT NOT NULL,
            description TEXT,
            preconditions TEXT,
            test_steps TEXT,
            expected_result TEXT,
            priority TEXT,
            test_type TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        return None

def extract_text_from_docx(file_path):
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        print(f"Error extracting DOCX: {e}")
        return None

def extract_text_from_txt(file_path):
    """Extract text from TXT file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        print(f"Error extracting TXT: {e}")
        return None

def extract_text_from_file(file_path, file_extension):
    """Extract text based on file type"""
    if file_extension == 'pdf':
        return extract_text_from_pdf(file_path)
    elif file_extension == 'docx':
        return extract_text_from_docx(file_path)
    elif file_extension == 'txt':
        return extract_text_from_txt(file_path)
    return None

def generate_test_cases_with_gemini(requirement_text):
    """Use Gemini AI to generate test cases from requirements"""
    try:
        model = genai.GenerativeModel('gemini-2.0-flash')
        
        prompt = f"""
You are an expert QA engineer with 10+ years of experience. Analyze the following requirement document THOROUGHLY and generate COMPREHENSIVE test cases covering ALL functionality mentioned in the document.

REQUIREMENT DOCUMENT:
{requirement_text}

CRITICAL INSTRUCTIONS:
1. READ THE ENTIRE DOCUMENT carefully and identify EVERY feature, function, and requirement
2. For EACH feature/requirement, generate test cases covering:
   - POSITIVE scenarios (happy path - things that should work)
   - NEGATIVE scenarios (error handling - things that should fail gracefully)
   - EDGE CASES (boundary conditions, unusual inputs, extreme values)
   - INTEGRATION scenarios (how features work together)

3. Generate detailed test cases in JSON format. Each test case MUST include:
   - test_case_name: A clear, specific, descriptive name
   - description: Brief description of what is being tested and why
   - preconditions: Any setup or conditions needed before testing
   - test_steps: Detailed step-by-step instructions (as a numbered list with \\n separators)
   - expected_result: What should happen if the test passes
   - priority: High, Medium, or Low (based on business impact)
   - test_type: One of the following categories

TEST COVERAGE REQUIREMENTS:

A. FUNCTIONAL TEST CASES (50% of total):
   
   For EACH feature in the document, create:
   
   1. POSITIVE Test Cases (Happy Path):
      - Valid inputs with expected successful outcomes
      - Standard user workflows
      - Normal operating conditions
      
   2. NEGATIVE Test Cases (Error Handling):
      - Invalid inputs (wrong format, type, length)
      - Missing required fields
      - Unauthorized access attempts
      - Invalid state transitions
      - Duplicate entries
      - Expired/invalid tokens or sessions
      
   3. EDGE CASES (Boundary Conditions):
      - Minimum and maximum values
      - Empty fields
      - Special characters
      - Very long inputs
      - Zero values
      - Null values
      - Concurrent operations
      
   4. INTEGRATION Test Cases:
      - How features interact with each other
      - Data flow between components
      - End-to-end workflows

B. NON-FUNCTIONAL TEST CASES (50% of total):
   
   1. PERFORMANCE Testing:
      - Response time under normal load
      - Response time under heavy load
      - Concurrent user handling
      - Database query performance
      - API response times
      - Page load times
      - Scalability testing
      
   2. SECURITY Testing:
      - Authentication mechanisms
      - Authorization and access control
      - SQL injection prevention
      - XSS (Cross-Site Scripting) prevention
      - CSRF protection
      - Password encryption
      - Session management
      - Data encryption in transit (HTTPS)
      - Input validation and sanitization
      - Brute force attack prevention
      
   3. USABILITY Testing:
      - User interface intuitiveness
      - Error message clarity
      - Navigation ease
      - Accessibility (WCAG compliance)
      - Mobile responsiveness
      - Form validation feedback
      
   4. RELIABILITY Testing:
      - Error recovery mechanisms
      - Data integrity validation
      - System stability under stress
      - Graceful degradation
      - Failover scenarios
      
   5. COMPATIBILITY Testing:
      - Browser compatibility (Chrome, Firefox, Safari, Edge)
      - Mobile device compatibility (iOS, Android)
      - Operating system compatibility
      - Screen resolution compatibility
      - Different network conditions
      
   6. MAINTAINABILITY Testing:
      - Code quality checks
      - Documentation completeness
      - Logging and monitoring
      - Error tracking

IMPORTANT REQUIREMENTS:
- Generate AT LEAST 20-30 test cases (more for complex documents)
- Ensure EVERY feature/requirement in the document has test coverage
- Maintain a 50-50 balance between Functional and Non-Functional tests
- Prioritize test cases based on business impact and risk
- Make test steps SPECIFIC and ACTIONABLE
- Include REALISTIC test data in examples

QUALITY CHECKLIST:
✓ Have you covered ALL features mentioned in the document?
✓ Does each feature have positive, negative, and edge case tests?
✓ Are security vulnerabilities tested (SQL injection, XSS, etc.)?
✓ Are performance requirements validated?
✓ Are all user workflows covered end-to-end?
✓ Are error messages and validation tested?
✓ Is the test data realistic and specific?

Return ONLY a valid JSON array of test cases.

Example format:
[
  {{
    "test_case_name": "Verify successful user login with valid credentials",
    "description": "Test that users can successfully log in using correct email and password",
    "preconditions": "1. User account exists in database\\n2. User is not already logged in\\n3. Account is not locked",
    "test_steps": "1. Navigate to login page (https://app.example.com/login)\\n2. Enter valid email: test@example.com\\n3. Enter valid password: Test@123456\\n4. Click 'Login' button\\n5. Verify redirect to dashboard",
    "expected_result": "User is successfully authenticated, session is created, and user is redirected to dashboard with welcome message displaying their name",
    "priority": "High",
    "test_type": "Functional"
  }},
  {{
    "test_case_name": "Verify login fails with invalid password",
    "description": "Test that login is rejected when user enters incorrect password",
    "preconditions": "1. User account exists in database\\n2. User is not logged in",
    "test_steps": "1. Navigate to login page\\n2. Enter valid email: test@example.com\\n3. Enter invalid password: WrongPassword123\\n4. Click 'Login' button\\n5. Observe error message",
    "expected_result": "Login fails, error message 'Invalid email or password' is displayed, user remains on login page, no session is created",
    "priority": "High",
    "test_type": "Functional"
  }},
  {{
    "test_case_name": "Verify SQL injection prevention in login form",
    "description": "Test that the login form properly sanitizes input and prevents SQL injection attacks",
    "preconditions": "Application is running and accessible",
    "test_steps": "1. Navigate to login page\\n2. Enter SQL injection payload in email field: admin' OR '1'='1' --\\n3. Enter any password: test123\\n4. Click 'Login' button\\n5. Verify system response",
    "expected_result": "Login fails, malicious input is sanitized/rejected, error message is displayed, no unauthorized access is granted, attempt is logged",
    "priority": "High",
    "test_type": "Security"
  }},
  {{
    "test_case_name": "Verify login page loads within acceptable time under normal load",
    "description": "Test that login page response time meets performance requirements with 100 concurrent users",
    "preconditions": "1. Application is deployed in production environment\\n2. Performance testing tool is configured\\n3. 100 concurrent users are simulated",
    "test_steps": "1. Configure load testing tool (JMeter/LoadRunner) for 100 concurrent users\\n2. Execute load test for login page\\n3. Measure average response time\\n4. Measure 95th percentile response time\\n5. Record any errors or timeouts",
    "expected_result": "Average response time is under 2 seconds, 95th percentile is under 3 seconds, no errors or timeouts occur, all users can access the page",
    "priority": "High",
    "test_type": "Performance"
  }}
]
"""
        
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        response_text = response.text.strip()
        
        # Parse JSON response
        test_cases = json.loads(response_text)
        return test_cases
    
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {e}")
        print(f"Response text: {response.text if 'response' in locals() else 'No response'}")
        return None
    except Exception as e:
        print(f"Error generating test cases: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and generate test cases"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only PDF, DOCX, and TXT files are allowed'}), 400
    
    try:
        # Save file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Extract text from file
        file_extension = filename.rsplit('.', 1)[1].lower()
        requirement_text = extract_text_from_file(file_path, file_extension)
        
        if not requirement_text:
            return jsonify({'error': 'Failed to extract text from file'}), 500
        
        # Generate test cases using Gemini AI
        test_cases = generate_test_cases_with_gemini(requirement_text)
        
        if not test_cases:
            return jsonify({'error': 'Failed to generate test cases'}), 500
        
        # Connect to database
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        # DELETE PREVIOUS TEST CASES FOR THIS FILE
        cursor.execute('DELETE FROM test_cases WHERE requirement_file = ?', (filename,))
        deleted_count = cursor.rowcount
        
        if deleted_count > 0:
            print(f"Deleted {deleted_count} previous test cases for file: {filename}")
        
        # Save new test cases to database
        saved_test_cases = []
        for tc in test_cases:
            # Handle list fields if Gemini returns them as arrays
            test_steps = tc.get('test_steps', '')
            if isinstance(test_steps, list):
                test_steps = '\n'.join(test_steps)
                
            preconditions = tc.get('preconditions', '')
            if isinstance(preconditions, list):
                preconditions = '\n'.join(preconditions)

            cursor.execute('''
                INSERT INTO test_cases 
                (requirement_file, test_case_name, description, preconditions, test_steps, expected_result, priority, test_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                filename,
                tc.get('test_case_name', ''),
                tc.get('description', ''),
                preconditions,
                test_steps,
                tc.get('expected_result', ''),
                tc.get('priority', 'Medium'),
                tc.get('test_type', 'Functional')
            ))
            
            test_case_id = cursor.lastrowid
            saved_test_cases.append({
                'id': test_case_id,
                'requirement_file': filename,
                **tc
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Successfully generated {len(saved_test_cases)} test cases for {filename}',
            'filename': filename,
            'test_cases': saved_test_cases,
            'replaced': deleted_count > 0
        }), 200
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases', methods=['GET'])
def get_test_cases():
    """Get all test cases or filter by filename"""
    try:
        # Get optional filename parameter
        filename = request.args.get('filename', None)
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Filter by filename if provided
        if filename:
            cursor.execute('SELECT * FROM test_cases WHERE requirement_file = ? ORDER BY created_at DESC', (filename,))
        else:
            cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        test_cases = []
        for row in rows:
            test_cases.append({
                'id': row['id'],
                'requirement_file': row['requirement_file'],
                'test_case_name': row['test_case_name'],
                'description': row['description'],
                'preconditions': row['preconditions'],
                'test_steps': row['test_steps'],
                'expected_result': row['expected_result'],
                'priority': row['priority'],
                'test_type': row['test_type'],
                'created_at': row['created_at']
            })
        
        conn.close()
        return jsonify(test_cases), 200
    
    except Exception as e:
        print(f"Error fetching test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['PUT'])
def update_test_case(test_case_id):
    """Update a test case"""
    try:
        data = request.json
        
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE test_cases 
            SET test_case_name = ?, description = ?, preconditions = ?, 
                test_steps = ?, expected_result = ?, priority = ?, test_type = ?
            WHERE id = ?
        ''', (
            data.get('test_case_name'),
            data.get('description'),
            data.get('preconditions'),
            data.get('test_steps'),
            data.get('expected_result'),
            data.get('priority'),
            data.get('test_type'),
            test_case_id
        ))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case updated successfully'}), 200
    
    except Exception as e:
        print(f"Error updating test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/<int:test_case_id>', methods=['DELETE'])
def delete_test_case(test_case_id):
    """Delete a test case"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases WHERE id = ?', (test_case_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Test case deleted successfully'}), 200
    
    except Exception as e:
        print(f"Error deleting test case: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/test-cases/clear-all', methods=['DELETE'])
def clear_all_test_cases():
    """Delete all test cases"""
    try:
        conn = sqlite3.connect('database.db')
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM test_cases')
        deleted_count = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        return jsonify({'message': f'Successfully deleted {deleted_count} test cases'}), 200
    
    except Exception as e:
        print(f"Error clearing test cases: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export', methods=['GET'])
def export_test_cases():
    """Export all test cases as Excel file"""
    try:
        conn = sqlite3.connect('database.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM test_cases ORDER BY created_at DESC')
        rows = cursor.fetchall()
        
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers
        headers = ['ID', 'Test Case Name', 'Description', 'Preconditions', 
                   'Test Steps', 'Expected Result', 'Priority', 'Test Type', 
                   'Requirement File', 'Created At']
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for row_num, row in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=row['id'])
            ws.cell(row=row_num, column=2, value=row['test_case_name'])
            ws.cell(row=row_num, column=3, value=row['description'])
            ws.cell(row=row_num, column=4, value=row['preconditions'])
            ws.cell(row=row_num, column=5, value=row['test_steps'])
            ws.cell(row=row_num, column=6, value=row['expected_result'])
            ws.cell(row=row_num, column=7, value=row['priority'])
            ws.cell(row=row_num, column=8, value=row['test_type'])
            ws.cell(row=row_num, column=9, value=row['requirement_file'])
            ws.cell(row=row_num, column=10, value=row['created_at'])
            
            # Apply text wrapping for long content
            for col in [2, 3, 4, 5, 6]:
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        ws.column_dimensions['J'].width = 20
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Save Excel file
        export_filename = f'test_cases_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        export_path = os.path.join(app.config['UPLOAD_FOLDER'], export_filename)
        
        wb.save(export_path)
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
    
    except Exception as e:
        print(f"Error exporting test cases: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
